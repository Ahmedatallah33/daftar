"""اختبارات توليد PDF و Excel (أسماء عربية، أمان المسارات)."""
from app.services.student_service import create_student
from app.services.session_service import add_session, add_video
from app.services import pdf_service as pdf
from app.services import excel_service as excel


def test_safe_filename_keeps_arabic_drops_separators():
    assert pdf._safe_filename("معاذ علي") == "معاذ علي"
    # محاولة الخروج من المجلد: الشرطات/النقاط تُزال
    cleaned = pdf._safe_filename("../../etc/passwd")
    assert "/" not in cleaned and ".." not in cleaned


def test_safe_filename_all_special_falls_back():
    assert pdf._safe_filename("///...") == "student"


def test_generate_invoice_arabic_name(tmp_path):
    st = create_student(name="محمد السوري", price_per_session=50, sessions_per_cycle=8)
    for _ in range(3):
        add_session(st.id)
    add_video(st.id, "فيديو")
    out = pdf.generate_invoice(st, amount=150.0, notes="شكراً")
    assert out.exists()
    assert out.stat().st_size > 0
    # لا يخرج عن مجلد الفواتير المؤقت
    assert out.parent == pdf.INVOICES_DIR


def test_generate_invoice_path_stays_in_dir():
    st = create_student(name="../../evil", price_per_session=10, sessions_per_cycle=8)
    add_session(st.id)
    out = pdf.generate_invoice(st, amount=10.0)
    assert out.parent == pdf.INVOICES_DIR


def test_export_students_xlsx():
    create_student(name="طالب ١", price_per_session=50)
    create_student(name="طالب ٢", price_per_session=60)
    out = excel.export_students_xlsx()
    assert out.exists()
    assert out.stat().st_size > 0

    from openpyxl import load_workbook
    wb = load_workbook(out)
    ws = wb.active
    assert ws.max_row == 3        # رأس + طالبان
    assert ws["A1"].value == "الاسم"
