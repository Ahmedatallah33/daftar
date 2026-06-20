"""Export students to xlsx — every saved field with Arabic headers."""
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.config import EXPORTS_DIR, WEEKDAY_AR
from app.services.student_service import (
    list_students, get_day_schedules, get_custom_fields,
    counted_sessions_map, counted_videos_map,
)


_HEADERS = [
    ("الاسم", 26),
    ("الهاتف", 18),
    ("هاتف ولي الأمر", 18),
    ("رابط Zoom", 36),
    ("اسم رابط Zoom", 22),
    ("رابط مجموعة واتساب", 36),
    ("سعر الحصة", 12),
    ("حصص/دورة", 11),
    ("الجدول الأسبوعي", 50),
    ("الوقت (الموروث)", 14),
    ("ملاحظات", 32),
    ("خانات إضافية", 50),
    ("نشط؟", 8),
    ("تاريخ الإنشاء", 18),
    ("الحصص المحتسبة", 14),
    ("الفيديوهات المحتسبة", 16),
]


def _format_schedules(student) -> str:
    sched = get_day_schedules(student)
    if not sched:
        return ""
    parts = []
    for code in WEEKDAY_AR:
        times = sched.get(code) or []
        if times:
            parts.append(f"{WEEKDAY_AR[code]}: {' ، '.join(times)}")
    return " | ".join(parts)


def _format_custom_fields(student) -> str:
    fields = get_custom_fields(student)
    if not fields:
        return ""
    return " | ".join(f"{f['label']}: {f['value']}" for f in fields)


def export_students_xlsx(out_path: Path | None = None) -> Path:
    """Export all students (active + inactive) to a single xlsx with all fields."""
    if out_path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = EXPORTS_DIR / f"students_{ts}.xlsx"
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "الطلاب"
    ws.sheet_view.rightToLeft = True  # RTL view in Excel

    # Header styling
    header_fill = PatternFill("solid", fgColor="1F3A8A")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    thin = Side(border_style="thin", color="9CA3AF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write headers
    for col_idx, (label, width) in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28

    # Data rows
    students = list_students(active_only=False)
    session_counts = counted_sessions_map()
    video_counts = counted_videos_map()
    for r_idx, s in enumerate(students, start=2):
        created = ""
        if s.created_at:
            try:
                created = s.created_at.strftime("%Y-%m-%d")
            except Exception:
                created = str(s.created_at)
        row_values = [
            s.name or "",
            s.phone or "",
            getattr(s, "parent_phone", "") or "",
            s.zoom_link or "",
            getattr(s, "zoom_link_name", "") or "",
            getattr(s, "whatsapp_group_link", "") or "",
            float(s.price_per_session or 0),
            int(s.sessions_per_cycle or 0),
            _format_schedules(s),
            s.session_time or "",
            s.notes or "",
            _format_custom_fields(s),
            "نعم" if s.is_active else "لا",
            created,
            session_counts.get(s.id, 0),
            video_counts.get(s.id, 0),
        ]
        for c_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(
                horizontal="right", vertical="center", wrap_text=True
            )
            cell.border = border
        ws.row_dimensions[r_idx].height = 22

    # Freeze header
    ws.freeze_panes = "A2"

    wb.save(str(out_path))
    return out_path
